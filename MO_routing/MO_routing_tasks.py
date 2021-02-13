# This code specifies the multi-objective drone routing problems and their single-objective VRP source(s)
# Permutation-encoded solutions are decoded for all target function evaluations
# A summary of the target and source problems is given in the file "MO_routing_tasks.pdf"

import math
import openpyxl




class M_151(object):
    def __init__(self):
        self.dim_source1 = 135
        self.dim_source2 = 140
        self.dim_target = 144
        self.pop_size_source = 100
        self.capacity_t = 20.0
        self.distance_limit_t = 50
        self.cx_rate = 0.75
        self.mut_rate = 0.2
        self.pf_target = 'pf_M_151.txt'
        self.number_of_sources = 2
        self.source_data1 = 'M_151_s1_data.xlsx'
        self.source_data2 = 'M_151_s2_data.xlsx'
        self.vec = [1, 1]


        # Source1 information
        source1_file = open("M_151_source1_instance.txt")
        source1_file.readline().split(" ")
        depot_s1 = source1_file.readline().split(" ")
        self.depot_s1_x = float(depot_s1[1])
        self.depot_s1_y = float(depot_s1[2])

        source1_x = []
        source1_y = []
        for i in range(self.dim_source1):
            source1_data = source1_file.readline().split(" ")
            source1_x.append(float(source1_data[1]))
            source1_y.append(float(source1_data[2]))

        source1_instance_x = []
        source1_instance_x.append(self.depot_s1_x)
        source1_instance_y = []
        source1_instance_y.append(self.depot_s1_y)
        for i in range(self.dim_source1):
            source1_instance_x.append(source1_x[i])
            source1_instance_y.append(source1_y[i])
        self.x_s1 = source1_instance_x
        self.y_s1 = source1_instance_y

        wb1 = openpyxl.load_workbook(self.source_data1)
        sheet1 = wb1.active
        s1_data = []
        for i in range(self.pop_size_source):
            data1 = []
            for j in range(self.dim_source1):
                data1.append(int(sheet1.cell(i + 1, j + 1)._value))
            s1_data.append(data1)
        self.data_s1 = s1_data


        # Source2 information
        source2_file = open("M_151_source2_instance.txt")
        source2_file.readline().split(" ")
        depot_s2 = source2_file.readline().split(" ")
        self.depot_s2_x = float(depot_s2[1])
        self.depot_s2_y = float(depot_s2[2])

        source2_x = []
        source2_y = []
        for i in range(self.dim_source2):
            source2_data = source2_file.readline().split(" ")
            source2_x.append(float(source2_data[1]))
            source2_y.append(float(source2_data[2]))

        source2_instance_x = []
        source2_instance_x.append(self.depot_s2_x)
        source2_instance_y = []
        source2_instance_y.append(self.depot_s2_y)
        for i in range(self.dim_source2):
            source2_instance_x.append(source2_x[i])
            source2_instance_y.append(source2_y[i])
        self.x_s2 = source2_instance_x
        self.y_s2 = source2_instance_y

        wb2 = openpyxl.load_workbook(self.source_data2)
        sheet2 = wb2.active
        s2_data = []
        for i in range(self.pop_size_source):
            data2 = []
            for j in range(self.dim_source2):
                data2.append(int(sheet2.cell(i + 1, j + 1)._value))
            s2_data.append(data2)
        self.data_s2 = s2_data


        # Target information
        target_file = open("M_151_target_instance.txt")
        target_file.readline().split(" ")
        depot_t = target_file.readline().split(" ")
        self.depot_t_x = float(depot_t[1])
        self.depot_t_y = float(depot_t[2])

        target_x = []
        target_y = []
        target_demand = []
        target_distance_depot = []
        target_distance_matrix = [[0 for _ in range(self.dim_target)] for _ in range(self.dim_target)]
        for i in range(self.dim_target):
            target_data = target_file.readline().split(" ")
            target_x.append(float(target_data[1]))
            target_y.append(float(target_data[2]))
            target_demand.append(float(target_data[3]))
            target_distance_depot.append(math.sqrt((target_x[i] - self.depot_t_x) * (target_x[i] - self.depot_t_x) +
                                                    (target_y[i] - self.depot_t_y) * (target_y[i] - self.depot_t_y)))
        for i in range(self.dim_target):
            for j in range(self.dim_target):
                target_distance_matrix[i][j] = (math.sqrt((target_x[i] - target_x[j]) ** 2 + (target_y[i] - target_y[j]) ** 2))
        self.demand_t = target_demand
        self.distance_depot_t = target_distance_depot
        self.distance_matrix_t = target_distance_matrix
        target_instance_x = []
        target_instance_x.append(self.depot_t_x)
        target_instance_y = []
        target_instance_y.append(self.depot_t_y)
        for j in range(self.dim_target):
            target_instance_x.append(target_x[j])
            target_instance_y.append(target_y[j])
        self.x_t = target_instance_x
        self.y_t = target_instance_y


    def target_functions(self, solution):
        distance = []
        vehicles = 0
        i = 0
        j = 0
        k = 0
        count = 0
        first_node = 1
        load = 0
        cost = 0
        while (i < self.dim_target):
            if load <= (self.capacity_t - self.demand_t[solution[i]]) \
                    and cost <= (self.distance_limit_t - self.distance_matrix_t[solution[i - 1]][solution[i]]):
                if j == 0:  # new route
                    load = load + self.demand_t[solution[i]]
                    if first_node == 1:  # first route
                        cost = self.distance_depot_t[solution[i]]
                        distance.append(cost)
                        first_node = 0
                        i = i + 1
                        j = j + 1
                        k = k + 1
                        count = count + 1
                    else:
                        if count == self.dim_target - 1:  # final route
                            distance.append(distance[k - 1] + self.distance_depot_t[solution[i]] + self.distance_depot_t[solution[i]])
                            i = i + 1
                            vehicles += 1
                        else:
                            cost = self.distance_depot_t[solution[i]]
                            distance.append(distance[k - 1] + cost)
                            i = i + 1
                            j = j + 1
                            k = k + 1
                            count = count + 1
                else:  # ongoing route
                    load = load + self.demand_t[solution[i]]
                    if count == self.dim_target - 1:  # final route
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]] + self.distance_depot_t[solution[i]])
                        i = i + 1
                        vehicles += 1
                    else:
                        cost = cost + self.distance_matrix_t[solution[i - 1]][solution[i]]
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]])
                        i = i + 1
                        k = k + 1
                        count = count + 1
            else:  # problem constraint(s) violated, vehicle return to depot
                distance.append(distance[k - 1] + self.distance_depot_t[solution[i - 1]])
                vehicles += 1
                j = 0
                k = k + 1
                load = 0
                cost = 0
                continue

        return distance[len(distance) - 1], vehicles




class X_214(object):
    def __init__(self):
        self.dim_source1 = 205
        self.dim_source2 = 213
        self.dim_target = 200
        self.pop_size_source = 100
        self.capacity_t = 47.2
        self.distance_limit_t = 50
        self.cx_rate = 0.75
        self.mut_rate = 0.2
        self.pf_target = 'pf_X_214.txt'
        self.number_of_sources = 2
        self.source_data1 = 'X_214_s1_data.xlsx'
        self.source_data2 = 'X_214_s2_data.xlsx'
        self.vec = [1, -1]


        # Source1 information
        source1_file = open("X_214_source1_instance.txt")
        source1_file.readline().split(" ")
        depot_s1 = source1_file.readline().split(" ")
        self.depot_s1_x = float(depot_s1[1])
        self.depot_s1_y = float(depot_s1[2])

        source1_x = []
        source1_y = []
        for i in range(self.dim_source1):
            source1_data = source1_file.readline().split(" ")
            source1_x.append(float(source1_data[1]))
            source1_y.append(float(source1_data[2]))

        source1_instance_x = []
        source1_instance_x.append(self.depot_s1_x)
        source1_instance_y = []
        source1_instance_y.append(self.depot_s1_y)
        for i in range(self.dim_source1):
            source1_instance_x.append(source1_x[i])
            source1_instance_y.append(source1_y[i])
        self.x_s1 = source1_instance_x
        self.y_s1 = source1_instance_y

        wb1 = openpyxl.load_workbook(self.source_data1)
        sheet1 = wb1.active
        s1_data = []
        for i in range(self.pop_size_source):
            data1 = []
            for j in range(self.dim_source1):
                data1.append(int(sheet1.cell(i + 1, j + 1)._value))
            s1_data.append(data1)
        self.data_s1 = s1_data

        # Source2 information
        source2_file = open("X_214_source2_instance.txt")
        source2_file.readline().split(" ")
        depot_s2 = source2_file.readline().split(" ")
        self.depot_s2_x = float(depot_s2[1])
        self.depot_s2_y = float(depot_s2[2])

        source2_x = []
        source2_y = []
        for i in range(self.dim_source2):
            source2_data = source2_file.readline().split(" ")
            source2_x.append(float(source2_data[1]))
            source2_y.append(float(source2_data[2]))

        source2_instance_x = []
        source2_instance_x.append(self.depot_s2_x)
        source2_instance_y = []
        source2_instance_y.append(self.depot_s2_y)
        for i in range(self.dim_source2):
            source2_instance_x.append(source2_x[i])
            source2_instance_y.append(source2_y[i])
        self.x_s2 = source2_instance_x
        self.y_s2 = source2_instance_y

        wb2 = openpyxl.load_workbook(self.source_data2)
        sheet2 = wb2.active
        s2_data = []
        for i in range(self.pop_size_source):
            data2 = []
            for j in range(self.dim_source2):
                data2.append(int(sheet2.cell(i + 1, j + 1)._value))
            s2_data.append(data2)
        self.data_s2 = s2_data


        # Target information
        target_file = open("X_214_target_instance.txt")
        target_file.readline().split(" ")
        depot_t = target_file.readline().split(" ")
        self.depot_t_x = float(depot_t[1])
        self.depot_t_y = float(depot_t[2])

        target_x = []
        target_y = []
        target_demand = []
        target_distance_depot = []
        target_distance_matrix = [[0 for _ in range(self.dim_target)] for _ in range(self.dim_target)]
        for i in range(self.dim_target):
            target_data = target_file.readline().split(" ")
            target_x.append(float(target_data[1]))
            target_y.append(float(target_data[2]))
            target_demand.append(float(target_data[3]))
            target_distance_depot.append(math.sqrt((target_x[i] - self.depot_t_x) * (target_x[i] - self.depot_t_x) +
                                                    (target_y[i] - self.depot_t_y) * (target_y[i] - self.depot_t_y)))
        for i in range(self.dim_target):
            for j in range(self.dim_target):
                target_distance_matrix[i][j] = (math.sqrt((target_x[i] - target_x[j]) ** 2 + (target_y[i] - target_y[j]) ** 2))
        self.demand_t = target_demand
        self.distance_depot_t = target_distance_depot
        self.distance_matrix_t = target_distance_matrix
        target_instance_x = []
        target_instance_x.append(self.depot_t_x)
        target_instance_y = []
        target_instance_y.append(self.depot_t_y)
        for j in range(self.dim_target):
            target_instance_x.append(target_x[j])
            target_instance_y.append(target_y[j])
        self.x_t = target_instance_x
        self.y_t = target_instance_y


    def target_functions(self, solution):
        distance = []
        vehicles = 0
        i = 0
        j = 0
        k = 0
        count = 0
        first_node = 1
        load = 0
        cost = 0
        while (i < self.dim_target):
            if load <= (self.capacity_t - self.demand_t[solution[i]]) \
                    and cost <= (self.distance_limit_t - self.distance_matrix_t[solution[i - 1]][solution[i]]):
                if j == 0:  # new route
                    load = load + self.demand_t[solution[i]]
                    if first_node == 1:  # first route
                        cost = self.distance_depot_t[solution[i]]
                        distance.append(cost)
                        first_node = 0
                        i = i + 1
                        j = j + 1
                        k = k + 1
                        count = count + 1
                    else:
                        if count == self.dim_target - 1:  # final route
                            distance.append(
                                distance[k - 1] + self.distance_depot_t[solution[i]] + self.distance_depot_t[
                                    solution[i]])
                            i = i + 1
                            vehicles += 1
                        else:
                            cost = self.distance_depot_t[solution[i]]
                            distance.append(distance[k - 1] + cost)
                            i = i + 1
                            j = j + 1
                            k = k + 1
                            count = count + 1
                else:  # ongoing route
                    load = load + self.demand_t[solution[i]]
                    if count == self.dim_target - 1:  # final route
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]] +
                                        self.distance_depot_t[solution[i]])
                        i = i + 1
                        vehicles += 1
                    else:
                        cost = cost + self.distance_matrix_t[solution[i - 1]][solution[i]]
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]])
                        i = i + 1
                        k = k + 1
                        count = count + 1
            else:  # problem constraint(s) violated, vehicle return to depot
                distance.append(distance[k - 1] + self.distance_depot_t[solution[i - 1]])
                vehicles += 1
                j = 0
                k = k + 1
                load = 0
                cost = 0
                continue

        return distance[len(distance) - 1], vehicles




class Golden_17(object):
    def __init__(self):
        self.dim_source1 = 210
        self.dim_source2 = 240
        self.dim_target = 222
        self.pop_size_source = 100
        self.capacity_t = 20.0
        self.distance_limit_t = 50
        self.cx_rate = 0.75
        self.mut_rate = 0.2
        self.pf_target = 'pf_Golden_17.txt'
        self.number_of_sources = 2
        self.source_data1 = 'Golden_17_s1_data.xlsx'
        self.source_data2 = 'Golden_17_s2_data.xlsx'
        self.vec = [1, 1]


        # Source1 information
        source1_file = open("Golden_17_source1_instance.txt")
        source1_file.readline().split(" ")
        depot_s1 = source1_file.readline().split(" ")
        self.depot_s1_x = float(depot_s1[1])
        self.depot_s1_y = float(depot_s1[2])

        source1_x = []
        source1_y = []
        for i in range(self.dim_source1):
            source1_data = source1_file.readline().split(" ")
            source1_x.append(float(source1_data[1]))
            source1_y.append(float(source1_data[2]))

        source1_instance_x = []
        source1_instance_x.append(self.depot_s1_x)
        source1_instance_y = []
        source1_instance_y.append(self.depot_s1_y)
        for i in range(self.dim_source1):
            source1_instance_x.append(source1_x[i])
            source1_instance_y.append(source1_y[i])
        self.x_s1 = source1_instance_x
        self.y_s1 = source1_instance_y

        wb1 = openpyxl.load_workbook(self.source_data1)
        sheet1 = wb1.active
        s1_data = []
        for i in range(self.pop_size_source):
            data1 = []
            for j in range(self.dim_source1):
                data1.append(int(sheet1.cell(i + 1, j + 1)._value))
            s1_data.append(data1)
        self.data_s1 = s1_data


        # Source2 information
        source2_file = open("Golden_17_source2_instance.txt")
        source2_file.readline().split(" ")
        depot_s2 = source2_file.readline().split(" ")
        self.depot_s2_x = float(depot_s2[1])
        self.depot_s2_y = float(depot_s2[2])

        source2_x = []
        source2_y = []
        for i in range(self.dim_source2):
            source2_data = source2_file.readline().split(" ")
            source2_x.append(float(source2_data[1]))
            source2_y.append(float(source2_data[2]))

        source2_instance_x = []
        source2_instance_x.append(self.depot_s2_x)
        source2_instance_y = []
        source2_instance_y.append(self.depot_s2_y)
        for i in range(self.dim_source2):
            source2_instance_x.append(source2_x[i])
            source2_instance_y.append(source2_y[i])
        self.x_s2 = source2_instance_x
        self.y_s2 = source2_instance_y

        wb2 = openpyxl.load_workbook(self.source_data2)
        sheet2 = wb2.active
        s2_data = []
        for i in range(self.pop_size_source):
            data2 = []
            for j in range(self.dim_source2):
                data2.append(int(sheet2.cell(i + 1, j + 1)._value))
            s2_data.append(data2)
        self.data_s2 = s2_data


        # Target information
        target_file = open("Golden_17_target_instance.txt")
        target_file.readline().split(" ")
        depot_t = target_file.readline().split(" ")
        self.depot_t_x = float(depot_t[1])
        self.depot_t_y = float(depot_t[2])

        target_x = []
        target_y = []
        target_demand = []
        target_distance_depot = []
        target_distance_matrix = [[0 for _ in range(self.dim_target)] for _ in range(self.dim_target)]
        for i in range(self.dim_target):
            target_data = target_file.readline().split(" ")
            target_x.append(float(target_data[1]))
            target_y.append(float(target_data[2]))
            target_demand.append(float(target_data[3]))
            target_distance_depot.append(math.sqrt((target_x[i] - self.depot_t_x) * (target_x[i] - self.depot_t_x) +
                                                    (target_y[i] - self.depot_t_y) * (target_y[i] - self.depot_t_y)))
        for i in range(self.dim_target):
            for j in range(self.dim_target):
                target_distance_matrix[i][j] = (math.sqrt((target_x[i] - target_x[j]) ** 2 + (target_y[i] - target_y[j]) ** 2))
        self.demand_t = target_demand
        self.distance_depot_t = target_distance_depot
        self.distance_matrix_t = target_distance_matrix
        target_instance_x = []
        target_instance_x.append(self.depot_t_x)
        target_instance_y = []
        target_instance_y.append(self.depot_t_y)
        for j in range(self.dim_target):
            target_instance_x.append(target_x[j])
            target_instance_y.append(target_y[j])
        self.x_t = target_instance_x
        self.y_t = target_instance_y


    def target_functions(self, solution):
        distance = []
        vehicles = 0
        i = 0
        j = 0
        k = 0
        count = 0
        first_node = 1
        load = 0
        cost = 0
        while (i < self.dim_target):
            if load <= (self.capacity_t - self.demand_t[solution[i]]) \
                    and cost <= (self.distance_limit_t - self.distance_matrix_t[solution[i - 1]][solution[i]]):
                if j == 0:  # new route
                    load = load + self.demand_t[solution[i]]
                    if first_node == 1:  # first route
                        cost = self.distance_depot_t[solution[i]]
                        distance.append(cost)
                        first_node = 0
                        i = i + 1
                        j = j + 1
                        k = k + 1
                        count = count + 1
                    else:
                        if count == self.dim_target - 1:  # final route
                            distance.append(
                                distance[k - 1] + self.distance_depot_t[solution[i]] + self.distance_depot_t[
                                    solution[i]])
                            i = i + 1
                            vehicles += 1
                        else:
                            cost = self.distance_depot_t[solution[i]]
                            distance.append(distance[k - 1] + cost)
                            i = i + 1
                            j = j + 1
                            k = k + 1
                            count = count + 1
                else:  # ongoing route
                    load = load + self.demand_t[solution[i]]
                    if count == self.dim_target - 1:  # final route
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]] +
                                        self.distance_depot_t[solution[i]])
                        i = i + 1
                        vehicles += 1
                    else:
                        cost = cost + self.distance_matrix_t[solution[i - 1]][solution[i]]
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]])
                        i = i + 1
                        k = k + 1
                        count = count + 1
            else:  # problem constraint(s) violated, vehicle return to depot
                distance.append(distance[k - 1] + self.distance_depot_t[solution[i - 1]])
                vehicles += 1
                j = 0
                k = k + 1
                load = 0
                cost = 0
                continue

        return distance[len(distance) - 1], vehicles




class Golden_18(object):
    def __init__(self):
        self.dim_source1 = 290
        self.dim_source2 = 300
        self.dim_target = 282
        self.pop_size_source = 100
        self.capacity_t = 20.0
        self.distance_limit_t = 50
        self.cx_rate = 0.75
        self.mut_rate = 0.2
        self.pf_target = 'pf_Golden_18.txt'
        self.number_of_sources = 2
        self.source_data1 = 'Golden_18_s1_data.xlsx'
        self.source_data2 = 'Golden_18_s2_data.xlsx'
        self.vec = [1, 1]


        # Source1 information
        source1_file = open("Golden_18_source1_instance.txt")
        source1_file.readline().split(" ")
        depot_s1 = source1_file.readline().split(" ")
        self.depot_s1_x = float(depot_s1[1])
        self.depot_s1_y = float(depot_s1[2])

        source1_x = []
        source1_y = []
        for i in range(self.dim_source1):
            source1_data = source1_file.readline().split(" ")
            source1_x.append(float(source1_data[1]))
            source1_y.append(float(source1_data[2]))

        source1_instance_x = []
        source1_instance_x.append(self.depot_s1_x)
        source1_instance_y = []
        source1_instance_y.append(self.depot_s1_y)
        for i in range(self.dim_source1):
            source1_instance_x.append(source1_x[i])
            source1_instance_y.append(source1_y[i])
        self.x_s1 = source1_instance_x
        self.y_s1 = source1_instance_y

        wb1 = openpyxl.load_workbook(self.source_data1)
        sheet1 = wb1.active
        s1_data = []
        for i in range(self.pop_size_source):
            data1 = []
            for j in range(self.dim_source1):
                data1.append(int(sheet1.cell(i + 1, j + 1)._value))
            s1_data.append(data1)
        self.data_s1 = s1_data


        # Source2 information
        source2_file = open("Golden_18_source2_instance.txt")
        source2_file.readline().split(" ")
        depot_s2 = source2_file.readline().split(" ")
        self.depot_s2_x = float(depot_s2[1])
        self.depot_s2_y = float(depot_s2[2])

        source2_x = []
        source2_y = []
        for i in range(self.dim_source2):
            source2_data = source2_file.readline().split(" ")
            source2_x.append(float(source2_data[1]))
            source2_y.append(float(source2_data[2]))

        source2_instance_x = []
        source2_instance_x.append(self.depot_s2_x)
        source2_instance_y = []
        source2_instance_y.append(self.depot_s2_y)
        for i in range(self.dim_source2):
            source2_instance_x.append(source2_x[i])
            source2_instance_y.append(source2_y[i])
        self.x_s2 = source2_instance_x
        self.y_s2 = source2_instance_y

        wb2 = openpyxl.load_workbook(self.source_data2)
        sheet2 = wb2.active
        s2_data = []
        for i in range(self.pop_size_source):
            data2 = []
            for j in range(self.dim_source2):
                data2.append(int(sheet2.cell(i + 1, j + 1)._value))
            s2_data.append(data2)
        self.data_s2 = s2_data


        # Target information
        target_file = open("Golden_18_target_instance.txt")
        target_file.readline().split(" ")
        depot_t = target_file.readline().split(" ")
        self.depot_t_x = float(depot_t[1])
        self.depot_t_y = float(depot_t[2])

        target_x = []
        target_y = []
        target_demand = []
        target_distance_depot = []
        target_distance_matrix = [[0 for _ in range(self.dim_target)] for _ in range(self.dim_target)]
        for i in range(self.dim_target):
            target_data = target_file.readline().split(" ")
            target_x.append(float(target_data[1]))
            target_y.append(float(target_data[2]))
            target_demand.append(float(target_data[3]))
            target_distance_depot.append(math.sqrt((target_x[i] - self.depot_t_x) * (target_x[i] - self.depot_t_x) +
                                                    (target_y[i] - self.depot_t_y) * (target_y[i] - self.depot_t_y)))
        for i in range(self.dim_target):
            for j in range(self.dim_target):
                target_distance_matrix[i][j] = (math.sqrt((target_x[i] - target_x[j]) ** 2 + (target_y[i] - target_y[j]) ** 2))
        self.demand_t = target_demand
        self.distance_depot_t = target_distance_depot
        self.distance_matrix_t = target_distance_matrix
        target_instance_x = []
        target_instance_x.append(self.depot_t_x)
        target_instance_y = []
        target_instance_y.append(self.depot_t_y)
        for j in range(self.dim_target):
            target_instance_x.append(target_x[j])
            target_instance_y.append(target_y[j])
        self.x_t = target_instance_x
        self.y_t = target_instance_y


    def target_functions(self, solution):
        distance = []
        vehicles = 0
        i = 0
        j = 0
        k = 0
        count = 0
        first_node = 1
        load = 0
        cost = 0
        while (i < self.dim_target):
            if load <= (self.capacity_t - self.demand_t[solution[i]]) \
                    and cost <= (self.distance_limit_t - self.distance_matrix_t[solution[i - 1]][solution[i]]):
                if j == 0:  # new route
                    load = load + self.demand_t[solution[i]]
                    if first_node == 1:  # first route
                        cost = self.distance_depot_t[solution[i]]
                        distance.append(cost)
                        first_node = 0
                        i = i + 1
                        j = j + 1
                        k = k + 1
                        count = count + 1
                    else:
                        if count == self.dim_target - 1:  # final route
                            distance.append(
                                distance[k - 1] + self.distance_depot_t[solution[i]] + self.distance_depot_t[
                                    solution[i]])
                            i = i + 1
                            vehicles += 1
                        else:
                            cost = self.distance_depot_t[solution[i]]
                            distance.append(distance[k - 1] + cost)
                            i = i + 1
                            j = j + 1
                            k = k + 1
                            count = count + 1
                else:  # ongoing route
                    load = load + self.demand_t[solution[i]]
                    if count == self.dim_target - 1:  # final route
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]] +
                                        self.distance_depot_t[solution[i]])
                        i = i + 1
                        vehicles += 1
                    else:
                        cost = cost + self.distance_matrix_t[solution[i - 1]][solution[i]]
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]])
                        i = i + 1
                        k = k + 1
                        count = count + 1
            else:  # problem constraint(s) violated, vehicle return to depot
                distance.append(distance[k - 1] + self.distance_depot_t[solution[i - 1]])
                vehicles += 1
                j = 0
                k = k + 1
                load = 0
                cost = 0
                continue

        return distance[len(distance) - 1], vehicles




class CMT4(object):
    def __init__(self):
        self.dim_source1 = 130
        self.dim_source2 = 140
        self.dim_source3 = 150
        self.dim_target = 133
        self.pop_size_source = 100
        self.capacity_t = 20.0
        self.distance_limit_t = 50
        self.cx_rate = 0.75
        self.mut_rate = 0.2
        self.pf_target = 'pf_vrpnc4.txt'
        self.number_of_sources = 3
        self.source_data1 = 'vrpnc4_s1_data.xlsx'
        self.source_data2 = 'vrpnc4_s2_data.xlsx'
        self.source_data3 = 'vrpnc4_s3_data.xlsx'
        self.vec = [1, 1]


        # Source1 information
        source1_file = open("vrpnc4_source1_instance.txt")
        source1_file.readline().split(" ")
        depot_s1 = source1_file.readline().split(" ")
        self.depot_s1_x = float(depot_s1[1])
        self.depot_s1_y = float(depot_s1[2])

        source1_x = []
        source1_y = []
        for i in range(self.dim_source1):
            source1_data = source1_file.readline().split(" ")
            source1_x.append(float(source1_data[1]))
            source1_y.append(float(source1_data[2]))

        source1_instance_x = []
        source1_instance_x.append(self.depot_s1_x)
        source1_instance_y = []
        source1_instance_y.append(self.depot_s1_y)
        for i in range(self.dim_source1):
            source1_instance_x.append(source1_x[i])
            source1_instance_y.append(source1_y[i])
        self.x_s1 = source1_instance_x
        self.y_s1 = source1_instance_y

        wb1 = openpyxl.load_workbook(self.source_data1)
        sheet1 = wb1.active
        s1_data = []
        for i in range(self.pop_size_source):
            data1 = []
            for j in range(self.dim_source1):
                data1.append(int(sheet1.cell(i + 1, j + 1)._value))
            s1_data.append(data1)
        self.data_s1 = s1_data


        # Source2 information
        source2_file = open("vrpnc4_source2_instance.txt")
        source2_file.readline().split(" ")
        depot_s2 = source2_file.readline().split(" ")
        self.depot_s2_x = float(depot_s2[1])
        self.depot_s2_y = float(depot_s2[2])

        source2_x = []
        source2_y = []
        for i in range(self.dim_source2):
            source2_data = source2_file.readline().split(" ")
            source2_x.append(float(source2_data[1]))
            source2_y.append(float(source2_data[2]))

        source2_instance_x = []
        source2_instance_x.append(self.depot_s2_x)
        source2_instance_y = []
        source2_instance_y.append(self.depot_s2_y)
        for i in range(self.dim_source2):
            source2_instance_x.append(source2_x[i])
            source2_instance_y.append(source2_y[i])
        self.x_s2 = source2_instance_x
        self.y_s2 = source2_instance_y

        wb2 = openpyxl.load_workbook(self.source_data2)
        sheet2 = wb2.active
        s2_data = []
        for i in range(self.pop_size_source):
            data2 = []
            for j in range(self.dim_source2):
                data2.append(int(sheet2.cell(i + 1, j + 1)._value))
            s2_data.append(data2)
        self.data_s2 = s2_data


        # Source3 information
        source3_file = open("vrpnc4_source3_instance.txt")
        source3_file.readline().split(" ")
        depot_s3 = source3_file.readline().split(" ")
        self.depot_s3_x = float(depot_s3[1])
        self.depot_s3_y = float(depot_s3[2])

        source3_x = []
        source3_y = []
        for i in range(self.dim_source3):
            source3_data = source3_file.readline().split(" ")
            source3_x.append(float(source3_data[1]))
            source3_y.append(float(source3_data[2]))

        source3_instance_x = []
        source3_instance_x.append(self.depot_s3_x)
        source3_instance_y = []
        source3_instance_y.append(self.depot_s3_y)
        for i in range(self.dim_source3):
            source3_instance_x.append(source3_x[i])
            source3_instance_y.append(source3_y[i])
        self.x_s3 = source3_instance_x
        self.y_s3 = source3_instance_y

        wb3 = openpyxl.load_workbook(self.source_data3)
        sheet3 = wb3.active
        s3_data = []
        for i in range(self.pop_size_source):
            data3 = []
            for j in range(self.dim_source3):
                data3.append(int(sheet3.cell(i + 1, j + 1)._value))
            s3_data.append(data3)
        self.data_s3 = s3_data


        # Target information
        target_file = open("vrpnc4_target_instance.txt")
        target_file.readline().split(" ")
        depot_t = target_file.readline().split(" ")
        self.depot_t_x = float(depot_t[1])
        self.depot_t_y = float(depot_t[2])

        target_x = []
        target_y = []
        target_demand = []
        target_distance_depot = []
        target_distance_matrix = [[0 for _ in range(self.dim_target)] for _ in range(self.dim_target)]
        for i in range(self.dim_target):
            target_data = target_file.readline().split(" ")
            target_x.append(float(target_data[1]))
            target_y.append(float(target_data[2]))
            target_demand.append(float(target_data[3]))
            target_distance_depot.append(math.sqrt((target_x[i] - self.depot_t_x) * (target_x[i] - self.depot_t_x) +
                                                    (target_y[i] - self.depot_t_y) * (target_y[i] - self.depot_t_y)))
        for i in range(self.dim_target):
            for j in range(self.dim_target):
                target_distance_matrix[i][j] = (math.sqrt((target_x[i] - target_x[j]) ** 2 + (target_y[i] - target_y[j]) ** 2))
        self.demand_t = target_demand
        self.distance_depot_t = target_distance_depot
        self.distance_matrix_t = target_distance_matrix
        target_instance_x = []
        target_instance_x.append(self.depot_t_x)
        target_instance_y = []
        target_instance_y.append(self.depot_t_y)
        for j in range(self.dim_target):
            target_instance_x.append(target_x[j])
            target_instance_y.append(target_y[j])
        self.x_t = target_instance_x
        self.y_t = target_instance_y


    def target_functions(self, solution):
        distance = []
        vehicles = 0
        i = 0
        j = 0
        k = 0
        count = 0
        first_node = 1
        load = 0
        cost = 0
        while (i < self.dim_target):
            if load <= (self.capacity_t - self.demand_t[solution[i]]) \
                    and cost <= (self.distance_limit_t - self.distance_matrix_t[solution[i - 1]][solution[i]]):
                if j == 0:  # new route
                    load = load + self.demand_t[solution[i]]
                    if first_node == 1:  # first route
                        cost = self.distance_depot_t[solution[i]]
                        distance.append(cost)
                        first_node = 0
                        i = i + 1
                        j = j + 1
                        k = k + 1
                        count = count + 1
                    else:
                        if count == self.dim_target - 1:  # final route
                            distance.append(
                                distance[k - 1] + self.distance_depot_t[solution[i]] + self.distance_depot_t[
                                    solution[i]])
                            i = i + 1
                            vehicles += 1
                        else:
                            cost = self.distance_depot_t[solution[i]]
                            distance.append(distance[k - 1] + cost)
                            i = i + 1
                            j = j + 1
                            k = k + 1
                            count = count + 1
                else:  # ongoing route
                    load = load + self.demand_t[solution[i]]
                    if count == self.dim_target - 1:  # final route
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]] +
                                        self.distance_depot_t[solution[i]])
                        i = i + 1
                        vehicles += 1
                    else:
                        cost = cost + self.distance_matrix_t[solution[i - 1]][solution[i]]
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]])
                        i = i + 1
                        k = k + 1
                        count = count + 1
            else:  # problem constraint(s) violated, vehicle return to depot
                distance.append(distance[k - 1] + self.distance_depot_t[solution[i - 1]])
                vehicles += 1
                j = 0
                k = k + 1
                load = 0
                cost = 0
                continue

        return distance[len(distance) - 1], vehicles




class X_162(object):
    def __init__(self):
        self.dim_source1 = 140
        self.dim_source2 = 150
        self.dim_source3 = 161
        self.dim_target = 143
        self.pop_size_source = 100
        self.capacity_t = 58.7
        self.distance_limit_t = 50
        self.cx_rate = 0.75
        self.mut_rate = 0.2
        self.pf_target = 'pf_X_162.txt'
        self.number_of_sources = 3
        self.source_data1 = 'X_162_s1_data.xlsx'
        self.source_data2 = 'X_162_s2_data.xlsx'
        self.source_data3 = 'X_162_s3_data.xlsx'
        self.vec = [1, 1]


        # Source1 information
        source1_file = open("X_162_source1_instance.txt")
        source1_file.readline().split(" ")
        depot_s1 = source1_file.readline().split(" ")
        self.depot_s1_x = float(depot_s1[1])
        self.depot_s1_y = float(depot_s1[2])

        source1_x = []
        source1_y = []
        for i in range(self.dim_source1):
            source1_data = source1_file.readline().split(" ")
            source1_x.append(float(source1_data[1]))
            source1_y.append(float(source1_data[2]))

        source1_instance_x = []
        source1_instance_x.append(self.depot_s1_x)
        source1_instance_y = []
        source1_instance_y.append(self.depot_s1_y)
        for i in range(self.dim_source1):
            source1_instance_x.append(source1_x[i])
            source1_instance_y.append(source1_y[i])
        self.x_s1 = source1_instance_x
        self.y_s1 = source1_instance_y

        wb1 = openpyxl.load_workbook(self.source_data1)
        sheet1 = wb1.active
        s1_data = []
        for i in range(self.pop_size_source):
            data1 = []
            for j in range(self.dim_source1):
                data1.append(int(sheet1.cell(i + 1, j + 1)._value))
            s1_data.append(data1)
        self.data_s1 = s1_data


        # Source2 information
        source2_file = open("X_162_source2_instance.txt")
        source2_file.readline().split(" ")
        depot_s2 = source2_file.readline().split(" ")
        self.depot_s2_x = float(depot_s2[1])
        self.depot_s2_y = float(depot_s2[2])

        source2_x = []
        source2_y = []
        for i in range(self.dim_source2):
            source2_data = source2_file.readline().split(" ")
            source2_x.append(float(source2_data[1]))
            source2_y.append(float(source2_data[2]))

        source2_instance_x = []
        source2_instance_x.append(self.depot_s2_x)
        source2_instance_y = []
        source2_instance_y.append(self.depot_s2_y)
        for i in range(self.dim_source2):
            source2_instance_x.append(source2_x[i])
            source2_instance_y.append(source2_y[i])
        self.x_s2 = source2_instance_x
        self.y_s2 = source2_instance_y

        wb2 = openpyxl.load_workbook(self.source_data2)
        sheet2 = wb2.active
        s2_data = []
        for i in range(self.pop_size_source):
            data2 = []
            for j in range(self.dim_source2):
                data2.append(int(sheet2.cell(i + 1, j + 1)._value))
            s2_data.append(data2)
        self.data_s2 = s2_data


        # Source3 information
        source3_file = open("X_162_source3_instance.txt")
        source3_file.readline().split(" ")
        depot_s3 = source3_file.readline().split(" ")
        self.depot_s3_x = float(depot_s3[1])
        self.depot_s3_y = float(depot_s3[2])

        source3_x = []
        source3_y = []
        for i in range(self.dim_source3):
            source3_data = source3_file.readline().split(" ")
            source3_x.append(float(source3_data[1]))
            source3_y.append(float(source3_data[2]))

        source3_instance_x = []
        source3_instance_x.append(self.depot_s3_x)
        source3_instance_y = []
        source3_instance_y.append(self.depot_s3_y)
        for i in range(self.dim_source3):
            source3_instance_x.append(source3_x[i])
            source3_instance_y.append(source3_y[i])
        self.x_s3 = source3_instance_x
        self.y_s3 = source3_instance_y

        wb3 = openpyxl.load_workbook(self.source_data3)
        sheet3 = wb3.active
        s3_data = []
        for i in range(self.pop_size_source):
            data3 = []
            for j in range(self.dim_source3):
                data3.append(int(sheet3.cell(i + 1, j + 1)._value))
            s3_data.append(data3)
        self.data_s3 = s3_data

        # Target information
        target_file = open("X_162_target_instance.txt")
        target_file.readline().split(" ")
        depot_t = target_file.readline().split(" ")
        self.depot_t_x = float(depot_t[1])
        self.depot_t_y = float(depot_t[2])

        target_x = []
        target_y = []
        target_demand = []
        target_distance_depot = []
        target_distance_matrix = [[0 for _ in range(self.dim_target)] for _ in range(self.dim_target)]
        for i in range(self.dim_target):
            target_data = target_file.readline().split(" ")
            target_x.append(float(target_data[1]))
            target_y.append(float(target_data[2]))
            target_demand.append(float(target_data[3]))
            target_distance_depot.append(math.sqrt((target_x[i] - self.depot_t_x) * (target_x[i] - self.depot_t_x) +
                                                    (target_y[i] - self.depot_t_y) * (target_y[i] - self.depot_t_y)))
        for i in range(self.dim_target):
            for j in range(self.dim_target):
                target_distance_matrix[i][j] = (math.sqrt((target_x[i] - target_x[j]) ** 2 + (target_y[i] - target_y[j]) ** 2))
        self.demand_t = target_demand
        self.distance_depot_t = target_distance_depot
        self.distance_matrix_t = target_distance_matrix
        target_instance_x = []
        target_instance_x.append(self.depot_t_x)
        target_instance_y = []
        target_instance_y.append(self.depot_t_y)
        for j in range(self.dim_target):
            target_instance_x.append(target_x[j])
            target_instance_y.append(target_y[j])
        self.x_t = target_instance_x
        self.y_t = target_instance_y


    def target_functions(self, solution):
        distance = []
        vehicles = 0
        i = 0
        j = 0
        k = 0
        count = 0
        first_node = 1
        load = 0
        cost = 0
        while (i < self.dim_target):
            if load <= (self.capacity_t - self.demand_t[solution[i]]) \
                    and cost <= (self.distance_limit_t - self.distance_matrix_t[solution[i - 1]][solution[i]]):
                if j == 0:  # new route
                    load = load + self.demand_t[solution[i]]
                    if first_node == 1:  # first route
                        cost = self.distance_depot_t[solution[i]]
                        distance.append(cost)
                        first_node = 0
                        i = i + 1
                        j = j + 1
                        k = k + 1
                        count = count + 1
                    else:
                        if count == self.dim_target - 1:  # final route
                            distance.append(
                                distance[k - 1] + self.distance_depot_t[solution[i]] + self.distance_depot_t[
                                    solution[i]])
                            i = i + 1
                            vehicles += 1
                        else:
                            cost = self.distance_depot_t[solution[i]]
                            distance.append(distance[k - 1] + cost)
                            i = i + 1
                            j = j + 1
                            k = k + 1
                            count = count + 1
                else:  # ongoing route
                    load = load + self.demand_t[solution[i]]
                    if count == self.dim_target - 1:  # final route
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]] +
                                        self.distance_depot_t[solution[i]])
                        i = i + 1
                        vehicles += 1
                    else:
                        cost = cost + self.distance_matrix_t[solution[i - 1]][solution[i]]
                        distance.append(distance[k - 1] + self.distance_matrix_t[solution[i - 1]][solution[i]])
                        i = i + 1
                        k = k + 1
                        count = count + 1
            else:  # problem constraint(s) violated, vehicle return to depot
                distance.append(distance[k - 1] + self.distance_depot_t[solution[i - 1]])
                vehicles += 1
                j = 0
                k = k + 1
                load = 0
                cost = 0
                continue

        return distance[len(distance) - 1], vehicles
